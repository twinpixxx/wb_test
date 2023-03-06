from openpyxl import load_workbook

def parse_xlsx(file) -> list[str]:
    wb = load_workbook(file)
    sheet = wb.active
    
    articles = []
    for row in range(1, sheet.max_row + 1):
        print(sheet.cell(row=row, column=1).value)
        articles.append(round(sheet.cell(row=row, column=1).value))

    return articles
